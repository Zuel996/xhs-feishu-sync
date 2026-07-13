"""CSV / Excel 导入采集器。

从小红书创作者中心导出的 CSV 或 Excel 文件中解析数据。
这是当浏览器采集和 API 都不可用时的降级方案。

支持的格式:
- 创作者中心: 笔记列表明细表.xlsx (Excel)
- 创作者中心: 笔记导出 CSV
- 新红/千瓜等第三方工具的标准导出 CSV
"""

import csv
import hashlib
import logging
import re
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from src.collectors.base import BaseCollector
from src.collectors.models import AccountProfile, CollectResult, NoteMetrics
from src.core.config import AccountInfo
from src.core.exceptions import CollectorError
from src.transformers.normalizer import parse_chinese_number

logger = logging.getLogger(__name__)

# ── CSV 列名映射（创作者中心 CSV 导出） ──

CSV_COLUMN_MAP = {
    "note_id": ["笔记ID", "note_id", "id"],
    "title": ["笔记标题", "标题", "title"],
    "note_type": ["笔记类型", "类型", "type"],
    "publish_date": ["发布时间", "发布日期", "publish_date", "date"],
    "url": ["笔记链接", "链接", "url", "link"],
    "views": ["阅读量", "浏览量", "浏览", "views", "阅读"],
    "likes": ["点赞数", "点赞", "likes", "like_count"],
    "favorites": ["收藏数", "收藏", "favorites", "fav_count"],
    "comments": ["评论数", "评论", "comments", "comment_count"],
    "shares": ["分享数", "分享", "shares", "share_count"],
    "followers": ["粉丝数", "粉丝", "followers", "follower_count"],
    "following": ["关注数", "关注", "following"],
    "total_likes": ["获赞数", "总获赞", "total_likes"],
    "total_collections": ["总收藏", "total_collections"],
}

# ── Excel 列名映射（创作者中心 Excel 导出） ──

EXCEL_COLUMN_MAP = {
    "title": ["笔记标题"],
    "note_type": ["体裁"],
    "publish_date": ["首次发布时间"],
    "impressions": ["曝光"],
    "views": ["观看量"],
    "ctr": ["封面点击率"],
    "likes": ["点赞"],
    "comments": ["评论"],
    "favorites": ["收藏"],
    "new_followers": ["涨粉"],
    "shares": ["分享"],
    "avg_watch_time": ["人均观看时长"],
    "danmaku": ["弹幕"],
}

# ── 笔记类型映射 ──

NOTE_TYPE_MAP = {
    "图文": "image",
    "图片": "image",
    "image": "image",
    "视频": "video",
    "video": "video",
    "短视频": "video",
}

# ── 支持的文件扩展名 ──

CSV_EXTENSIONS = {".csv"}
EXCEL_EXTENSIONS = {".xlsx", ".xls"}
ALL_EXTENSIONS = CSV_EXTENSIONS | EXCEL_EXTENSIONS


class CSVImportCollector(BaseCollector):
    """从 CSV / Excel 文件导入小红书数据。

    用法:
        collector = CSVImportCollector(base_dir="data/csv_imports")
        result = await collector.collect_all(account, target_date)
    """

    def __init__(self, base_dir: str = "data/csv_imports"):
        super().__init__()
        self.base_dir = Path(base_dir)
        self.base_dir.mkdir(parents=True, exist_ok=True)

    # ── 文件查找 ──

    def _find_data_file(self, account_id: str) -> Optional[Path]:
        """查找指定账号的最新数据文件（CSV 或 Excel）。

        按命名约定查找: {base_dir}/{account_id}/*
        优先返回最新修改的文件。
        """
        account_dir = self.base_dir / account_id
        if not account_dir.exists():
            return None
        files = sorted(
            [p for p in account_dir.iterdir()
             if p.is_file() and p.suffix.lower() in ALL_EXTENSIONS],
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        return files[0] if files else None

    # ── 表头映射 ──

    def _map_header(
        self, header: list[str], column_map: dict
    ) -> dict[str, int]:
        """将表头映射到内部字段名。返回 {internal_field: column_index}。"""
        mapping: dict[str, int] = {}
        for internal_name, candidate_names in column_map.items():
            for i, col in enumerate(header):
                if col.strip() in candidate_names:
                    mapping[internal_name] = i
                    break
        return mapping

    # ── 日期解析 ──

    def _parse_date(self, value: str) -> Optional[date]:
        """解析日期字符串，支持多种格式。

        支持:
        - 2026-07-10
        - 2026/07/10
        - 2026.07.10
        - 2026年07月10日17时35分27秒 (创作者中心 Excel 格式)
        - 2026年7月10日
        """
        if not value:
            return None
        value = str(value).strip()

        # Excel 中文格式: "2026年07月10日17时35分27秒"
        cn_match = re.match(
            r"(\d{4})年(\d{1,2})月(\d{1,2})日.*", value
        )
        if cn_match:
            return date(
                int(cn_match.group(1)),
                int(cn_match.group(2)),
                int(cn_match.group(3)),
            )

        # 标准格式
        formats = [
            "%Y-%m-%d",
            "%Y/%m/%d",
            "%Y.%m.%d",
            "%Y-%m-%d %H:%M:%S",
            "%m/%d/%Y",
        ]
        for fmt in formats:
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        try:
            return date.fromisoformat(value)
        except (ValueError, TypeError):
            pass
        return None

    # ── 通用行解析 ──

    def _get_val(
        self, mapping: dict[str, int], row: list[str],
        field: str, default: str = ""
    ) -> str:
        """从映射行中安全获取值。"""
        idx = mapping.get(field)
        if idx is not None and idx < len(row):
            return str(row[idx]).strip()
        return default

    def _parse_note_row(
        self, row: list[str], mapping: dict[str, int],
        account: AccountInfo,
    ) -> NoteMetrics:
        """将一行数据解析为 NoteMetrics。"""

        title = self._get_val(mapping, row, "title")
        note_type = NOTE_TYPE_MAP.get(
            self._get_val(mapping, row, "note_type", "图文").lower(),
            "image",
        )

        # note_id: 优先用 CSV 的笔记ID，否则用标题 MD5 生成
        note_id = self._get_val(mapping, row, "note_id")
        if not note_id:
            if title:
                id_source = f"{account.account_id}_{title}"
            else:
                # 无标题笔记：用发布日期+体裁+曝光+观看量组合生成唯一ID
                pub_date = self._get_val(mapping, row, "publish_date")
                note_type_val = self._get_val(mapping, row, "note_type")
                impressions_val = self._get_val(mapping, row, "impressions")
                views_val = self._get_val(mapping, row, "views")
                id_source = f"{account.account_id}_{pub_date}_{note_type_val}_{impressions_val}_{views_val}"
            note_id = hashlib.md5(id_source.encode()).hexdigest()[:16]

        # 处理浮点字段（CTR、人均观看时长）
        ctr_str = self._get_val(mapping, row, "ctr")
        try:
            ctr_val = float(ctr_str) if ctr_str else 0.0
        except ValueError:
            ctr_val = 0.0

        avg_watch_str = self._get_val(mapping, row, "avg_watch_time")
        try:
            avg_watch_val = float(avg_watch_str) if avg_watch_str else 0.0
        except ValueError:
            avg_watch_val = 0.0

        return NoteMetrics(
            note_id=note_id,
            account_id=account.account_id,
            title=title,
            note_type=note_type,
            publish_date=self._parse_date(
                self._get_val(mapping, row, "publish_date")
            ),
            url=self._get_val(mapping, row, "url"),
            impressions=parse_chinese_number(
                self._get_val(mapping, row, "impressions")
            ),
            views=parse_chinese_number(
                self._get_val(mapping, row, "views")
            ),
            likes=parse_chinese_number(
                self._get_val(mapping, row, "likes")
            ),
            favorites=parse_chinese_number(
                self._get_val(mapping, row, "favorites")
            ),
            comments=parse_chinese_number(
                self._get_val(mapping, row, "comments")
            ),
            shares=parse_chinese_number(
                self._get_val(mapping, row, "shares")
            ),
            ctr=ctr_val,
            avg_watch_time=avg_watch_val,
            new_followers=parse_chinese_number(
                self._get_val(mapping, row, "new_followers")
            ),
            danmaku=parse_chinese_number(
                self._get_val(mapping, row, "danmaku")
            ),
        )

    # ── 抽象方法实现 ──

    async def collect_account_profile(
        self, account: AccountInfo
    ) -> AccountProfile:
        """从数据文件提取账号概览数据。

        CSV 格式：从汇总行提取粉丝/关注/获赞等。
        Excel 格式：无账号汇总数据，返回基础 profile。
        """
        data_path = self._find_data_file(account.account_id)
        if not data_path:
            raise CollectorError(
                f"找不到账号 {account.account_id} 的数据文件。"
                f"请将 CSV/Excel 放入 {self.base_dir / account.account_id}/"
            )

        profile = AccountProfile(
            account_id=account.account_id,
            xhs_user_id=account.xhs_user_id,
            username=account.xhs_username,
            display_name=account.display_name,
            competitor=account.competitor,
            # 使用 accounts.yaml 的手动配置作为 fallback
            follower_count=account.follower_count,
            following_count=account.following_count,
            total_likes=account.total_likes,
            total_collections=account.total_collections,
        )

        ext = data_path.suffix.lower()

        if ext == ".csv":
            profile = self._parse_csv_account_profile(data_path, account, profile)
        elif ext in (".xlsx", ".xls"):
            # Excel 导出不含账号汇总，尝试从笔记数据推算
            logger.info(
                "Excel 文件不含账号汇总数据，使用基础 profile"
            )

        logger.info(
            "账号概览: %s (粉丝: %d)",
            account.display_name, profile.follower_count,
        )
        return profile

    def _parse_csv_account_profile(
        self, path: Path, account: AccountInfo,
        profile: AccountProfile,
    ) -> AccountProfile:
        """从 CSV 文件解析账号汇总数据。"""
        with open(path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            header_row = next(reader, None)
            if not header_row:
                return profile
            header = [h.strip() for h in header_row]
            mapping = self._map_header(header, CSV_COLUMN_MAP)

            for row in reader:
                note_id = self._get_val(mapping, row, "note_id")
                if not note_id or note_id.lower() in ("summary", "汇总", "合计"):
                    profile.follower_count = parse_chinese_number(
                        self._get_val(mapping, row, "followers", "0")
                    )
                    profile.following_count = parse_chinese_number(
                        self._get_val(mapping, row, "following", "0")
                    )
                    profile.total_likes = parse_chinese_number(
                        self._get_val(mapping, row, "total_likes", "0")
                    )
                    profile.total_collections = parse_chinese_number(
                        self._get_val(mapping, row, "total_collections", "0")
                    )
                    break
        return profile

    async def collect_notes_data(
        self, account: AccountInfo, target_date: Optional[date] = None
    ) -> list[NoteMetrics]:
        """从数据文件解析笔记数据。支持 CSV 和 Excel。"""
        data_path = self._find_data_file(account.account_id)
        if not data_path:
            raise CollectorError(
                f"找不到账号 {account.account_id} 的数据文件。"
            )

        ext = data_path.suffix.lower()

        if ext in (".xlsx", ".xls"):
            notes = self._parse_excel_notes(data_path, account, target_date)
        else:
            notes = self._parse_csv_notes(data_path, account, target_date)

        logger.info(
            "数据导入完成: %s — %d 篇笔记 (文件: %s)",
            account.display_name, len(notes), data_path.name,
        )
        return notes

    # ── CSV 解析 ──

    def _parse_csv_notes(
        self, path: Path, account: AccountInfo,
        target_date: Optional[date],
    ) -> list[NoteMetrics]:
        notes: list[NoteMetrics] = []
        with open(path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            header_row = next(reader, None)
            if not header_row:
                return notes
            header = [h.strip() for h in header_row]
            mapping = self._map_header(header, CSV_COLUMN_MAP)

            for row in reader:
                if not row or all(c.strip() == "" for c in row):
                    continue
                try:
                    note = self._parse_note_row(row, mapping, account)
                    if target_date and note.publish_date:
                        if note.publish_date != target_date:
                            continue
                    if not note.note_id or len(note.note_id) < 5:
                        continue
                    note.sort_order = len(notes) + 1
                    notes.append(note)
                except Exception as e:
                    logger.warning(
                        "解析行失败: %s (行内容: %s)", e, row[:3]
                    )
                    continue
        return notes

    # ── Excel 解析 ──

    def _parse_excel_notes(
        self, path: Path, account: AccountInfo,
        target_date: Optional[date],
    ) -> list[NoteMetrics]:
        """解析创作者中心 Excel 导出文件。

        Excel 结构:
        - 第1行: 文件说明（"最多导出排序后前1000条笔记"）
        - 第2行: 实际表头
        - 第3行起: 数据行
        """
        import openpyxl

        wb = openpyxl.load_workbook(str(path), data_only=True)
        ws = wb.active
        if ws.max_row < 3:
            wb.close()
            return []

        # 读取表头（第2行）
        header = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=2, column=col).value
            header.append(str(val).strip() if val else "")
        mapping = self._map_header(header, EXCEL_COLUMN_MAP)

        # 检查必要列
        if "title" not in mapping:
            wb.close()
            raise CollectorError(
                f"Excel 文件缺少必要列（笔记标题/体裁等）。"
                f"表头: {header}"
            )

        notes: list[NoteMetrics] = []
        for r in range(3, ws.max_row + 1):
            row = []
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                row.append(str(val).strip() if val is not None else "")

            # 跳过空行
            if all(v == "" for v in row):
                continue

            try:
                note = self._parse_note_row(row, mapping, account)

                if target_date and note.publish_date:
                    if note.publish_date != target_date:
                        continue

                note.sort_order = len(notes) + 1  # Excel行顺序
                notes.append(note)
            except Exception as e:
                logger.warning(
                    "解析 Excel 行失败: %s (行 %d: %s)",
                    e, r, row[:3],
                )
                continue

        wb.close()
        return notes

    async def validate_connection(self) -> bool:
        """检查数据目录是否存在。"""
        return self.base_dir.exists()
